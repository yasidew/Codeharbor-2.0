import csv
import json
import os
import re
from javalang import parse
import openpyxl
from datetime import datetime
import aiohttp
import asyncio
import pandas as pd
from django.conf import settings
from django.core.serializers import serialize
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from matplotlib import pyplot as plt
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from scipy.stats import pearsonr, spearmanr
import seaborn as sns
from transformers import RobertaTokenizer, RobertaForSequenceClassification
from Drinks.models import Drink
from analysis.code_analyzer import CodeAnalyzer
from analysis.java_code_analyser import JavaCodeAnalyzer
from analysis.javascript_code_analyser import JavaScriptCodeAnalyzer
from analysis.php_code_analyser import PHPCodeAnalyzer
from analysis.python_code_analyser import PythonCodeAnalyzer
from code_analysis.models import CodeSnippet, Project, CodeAnalysisHistory, JavaCodeSnippet, JavaProject
from .serializers import DrinkSerializer
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from .complexity_calculator import calculate_code_complexity_multiple_files, ai_recommend_refactoring, \
    calculate_code_complexity_by_method
from .complexity_calculator import calculate_code_complexity_line_by_line
from .metrics import analyze_code_complexity, load_guidelines, count_lines_of_code, count_functions_and_length, \
    count_duplicate_code_percentage, calculate_comment_density, calculate_readability_score, calculate_complexity_score, \
    categorize_value, find_duplicate_code, java_analyze_code_complexity, java_count_lines_of_code, \
    java_count_duplicate_code_percentage, java_calculate_comment_density, java_calculate_readability_score, \
     java_calculate_complexity_score, java_categorize_value, java_load_guidelines, \
    java_count_classes_and_methods, java_calculate_nesting_depth, java_find_duplicate_code

from prettytable import PrettyTable
import statsmodels.api as sm
from transformers import T5ForConditionalGeneration, AutoTokenizer
from rest_framework.decorators import api_view
import concurrent.futures
from functools import lru_cache
from dotenv import load_dotenv
import openai
import base64
from django.http import JsonResponse
from django.shortcuts import render
import json
import requests
from openai import OpenAI
import os
import torch
import ast
from transformers import AutoTokenizer, T5ForConditionalGeneration
from dotenv import load_dotenv
from django.shortcuts import render
from rest_framework.decorators import api_view
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from io import BytesIO
from .utils import calculate_loc, calculate_readability, extract_code_from_github
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime

load_dotenv()
# OpenAI API Client
# client = openai.OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

# Initialize model and tokenizer once
MODEL_PATH = "./models/custom_seq2seq_model"
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

# Global instances for reuse
tokenizer = AutoTokenizer.from_pretrained(MODEL_PATH)
model = T5ForConditionalGeneration.from_pretrained(MODEL_PATH).to(device)
model.eval()

# Optimize PyTorch performance
torch.backends.cudnn.benchmark = True
torch.backends.cuda.matmul.allow_tf32 = True

# FLASK_API_URL = "http://16.171.138.50:5000/predict"  # Update if deployed on AWS
FLASK_API_URL = "http://127.0.0.1:5000/predict"

def call_flask_model(snippet):
    """Calls Flask API and gets prediction from T5 model."""
    try:
        response = requests.post(FLASK_API_URL, json={"snippet": snippet}, timeout=15)
        if response.status_code == 200:
            return response.json().get("generated_text", "❌ No response from model")
        else:
            return f"❌ Flask API Error: {response.status_code} - {response.text}"
    except requests.exceptions.RequestException as e:
        return f"❌ Flask API Request Failed: {str(e)}"

# async def call_flask_model_async(snippet):
#     """Calls Flask API asynchronously and gets prediction from T5 model."""
#     async with aiohttp.ClientSession() as session:
#         try:
#             async with session.post(FLASK_API_URL, json={"snippet": snippet}, timeout=15) as response:
#                 if response.status == 200:
#                     return await response.json()
#                 else:
#                     return f"❌ Flask API Error: {response.status} - {await response.text()}"
#         except Exception as e:
#             return f"❌ Flask API Request Failed: {str(e)}"



def home(request):
    return render(request, 'home.html')


def refactor_view(request):
    # Your view logic here
    return render(request, 'refactor.html')


def upload_code(request):
    return render(request, 'upload_code.html')


def refactor_code(request):
    return render(request, 'refactor_code.html')


@api_view(['GET', 'POST'])
def python_code_analysis(request):
    recommendations = {"files": {}, "pasted_code": {}}  # Initialize as a dictionary
    summary = {
        "total_vulnerabilities": 0,
        "categories": {},
        "files_analyzed": 0,
    }
    code = ""

    if request.method == 'POST':
        # Handle pasted code
        code = request.POST.get('code', '').strip()
        if code:
            try:
                analyzer = PythonCodeAnalyzer(code)
                pasted_results = analyzer.generate_recommendations()
                grouped = group_recommendations_by_line(pasted_results)
                recommendations["pasted_code"] = {
                    "code": code,
                    "results": grouped,
                }
                # Update summary
                for recs in grouped.values():
                    for rec in recs:
                        summary["total_vulnerabilities"] += 1
                        summary["categories"].setdefault(rec["rule"], 0)
                        summary["categories"][rec["rule"]] += 1
            except Exception as e:
                recommendations["pasted_code"] = {
                    "error": f"Error analyzing pasted code: {str(e)}"
                }

        # Handle uploaded files
        files = request.FILES.getlist('files')
        summary["files_analyzed"] = len(files)
        if files:
            for file in files:
                try:
                    content = file.read().decode('utf-8')  # Assuming UTF-8 encoding
                    analyzer = PythonCodeAnalyzer(content)
                    file_results = analyzer.generate_recommendations()
                    grouped = group_recommendations_by_line(file_results)
                    recommendations["files"][file.name] = {
                        "content": content,
                        "results": grouped,
                    }
                    # Update summary
                    for recs in grouped.values():
                        for rec in recs:
                            summary["total_vulnerabilities"] += 1
                            summary["categories"].setdefault(rec["rule"], 0)
                            summary["categories"][rec["rule"]] += 1
                except Exception as e:
                    recommendations["files"][file.name] = {
                        "error": f"Error analyzing file: {str(e)}"
                    }

        return render(
            request,
            'python_code_analysis.html',
            {'recommendations': recommendations, 'summary': summary, 'code': code}
        )

    return render(request, 'python_code_analysis.html', {'code': code})


def group_recommendations_by_line(recommendations):
    """
    Group recommendations by line number.
    """
    grouped = {}
    for rec in recommendations:
        line = rec.get("line", "unknown")
        grouped.setdefault(line, []).append({
            "rule": rec.get("rule"),
            "message": rec.get("message"),
        })
    return grouped

################################ java ##############################

JAVA_MODEL_PATH = "./models/java_seq2seq_model"  # Update with the correct path
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

# ✅ Initialize model & tokenizer globally
java_tokenizer = AutoTokenizer.from_pretrained(JAVA_MODEL_PATH)
java_model = T5ForConditionalGeneration.from_pretrained(JAVA_MODEL_PATH).to(device)
java_model.eval()  # Set to evaluation mode for inference

def java_generate_suggestion(code_snippet):
    """
    Uses the Java-trained T5 model to generate AI-powered suggestions.
    """
    try:
        inputs = java_tokenizer(code_snippet, return_tensors="pt", truncation=True, padding="max_length", max_length=512)
        inputs = {k: v.to(device) for k, v in inputs.items()}

        with torch.no_grad():
            output = java_model.generate(**inputs, max_length=128)

        return java_tokenizer.decode(output[0], skip_special_tokens=True)

    except Exception as e:
        return f"❌ Error generating suggestion: {str(e)}"


import javalang

import re

def java_split_code_snippets(code):
    """
    Splits Java code into individual methods, logical units, and extracts vulnerable blocks separately.
    """
    try:
        # ✅ Normalize whitespace
        code = code.strip()

        # ✅ Step 1: Extract method definitions
        method_pattern = r'(public|private|protected|static|\s)+\s+\w+\s+\w+\s*\(.*?\)\s*\{(?:[^{}]*\{[^{}]*\}[^{}]*|[^{}]*)*\}'
        method_snippets = re.findall(method_pattern, code, re.DOTALL)

        # ✅ Step 2: Extract class declarations separately
        class_pattern = r'(public|private|protected)?\s*(class|interface|enum)\s+\w+\s*\{(?:[^{}]*\{[^{}]*\}[^{}]*|[^{}]*)*\}'
        class_snippets = re.findall(class_pattern, code, re.DOTALL)

        # ✅ Step 3: Extract **only important** logical breaks (Avoid too-large snippets)
        logical_breaks = [
            r'Statement stmt',      # SQL Injection
            r'Runtime\.getRuntime', # Command Injection
            r'new FileReader',      # Insecure File Handling
            r'session\.setAttribute', # Session Management
            r'BufferedReader',      # File Handling
            r'Cookie\(',            # Cookie issues
            r'return recursiveFunction' # Infinite Recursion
        ]

        final_snippets = []
        for snippet in method_snippets + class_snippets:
            logical_snippets = re.split("|".join(logical_breaks), snippet)
            for part in logical_snippets:
                cleaned_part = part.strip()
                if len(cleaned_part) > 5:  # Avoid empty snippets
                    final_snippets.append(cleaned_part)

        return final_snippets

    except Exception as e:
        print(f"❌ Error splitting Java snippets: {str(e)}")
        return []


def extract_java_block(code, start_line):
    """
    Extracts a Java code block (method/class) using balanced braces.
    """
    lines = code.splitlines()
    stack = []
    extracted = []
    inside_string = False
    escape_char = False

    for i in range(start_line, len(lines)):
        line = lines[i]
        extracted.append(line)

        for char in line:
            if char == '"' and not escape_char:
                inside_string = not inside_string
            escape_char = (char == '\\') if not escape_char else False

            if not inside_string:
                if char == '{':
                    stack.append('{')
                elif char == '}':
                    if stack:
                        stack.pop()
                    if not stack:
                        return "\n".join(extracted)

    return "\n".join(extracted)

def extract_logical_block(method_body, keyword):
    """
    Extracts a small logical block containing the keyword.
    """
    lines = method_body.split("\n")
    extracted = []

    for line in lines:
        if keyword in line:
            start_index = max(0, lines.index(line) - 1)
            end_index = min(len(lines), lines.index(line) + 2)
            extracted.extend(lines[start_index:end_index])

    return "\n".join(extracted) if extracted else None



@api_view(['GET', 'POST'])
def java_code_analysis(request):
    """
    Handles Java code analysis: Accepts pasted/uploaded/imported code, categorizes vulnerabilities,
    assigns severity levels, and stores results in PostgreSQL.
    """
    suggestions = []
    summary = {
        "total_snippets": 0,
        "total_suggestions": 0,
        "total_lines": 0,
        "categories": {},
        "severity": {"Critical": 0, "Medium": 0, "Low": 0},
        "complexity_metrics": {},
    }
    code_snippet = ""
    final_guideline = ""
    all_code_snippets = []  # ✅ Store all snippets (pasted, uploaded, GitHub)

    if request.method == 'POST':
        # ✅ Handle GitHub repository submission (JSON request)
        if request.content_type == 'application/json':
            data = json.loads(request.body)
            github_repo_url = data.get("github_url", "").strip()

            if github_repo_url:
                files, error = fetch_github_files(github_repo_url)
                if error:
                    return JsonResponse({"error": error})
                if files:
                    return JsonResponse({"files": files})  # ✅ Return GitHub files to frontend

        # ✅ Handle manually entered code & uploaded files
        code_snippet = request.POST.get('code', '').strip()
        project_name = request.POST.get('project_name', '').strip()

        # ✅ Auto-generate project name if none is provided
        if not project_name:
            project_name = f"JavaProject_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            print(f"🆕 Auto-generated project name: {project_name}")

        # ✅ Check if project exists, else create
        project, _ = JavaProject.objects.get_or_create(name=project_name)

        # ✅ Fetch uploaded files
        uploaded_files = request.FILES.getlist('files')

        # ✅ Process uploaded files
        if uploaded_files:
            for uploaded_file in uploaded_files:
                file_name = uploaded_file.name
                file_content = uploaded_file.read().decode('utf-8')
                all_code_snippets.append({"name": file_name, "code": file_content})

        # ✅ Add manually pasted code
        if code_snippet:
            all_code_snippets.append({"name": "Pasted Code", "code": code_snippet})

        # ✅ Handle GitHub repository files (if provided)
        github_repo_url = request.POST.get("github_url", "").strip()
        if github_repo_url:
            github_files, error = fetch_github_files(github_repo_url)
            if github_files:
                all_code_snippets.extend(github_files)

        if not all_code_snippets:
            return render(request, 'java_code_analysis.html', {
                "code": "", "suggestions": [], "summary": summary,"final_guideline": "",
                "error": "No Java code provided for analysis."
            })

        # ✅ Process each file/snippet
        for file_data in all_code_snippets:
            file_name = file_data["name"]
            file_code = file_data["code"]

            # ✅ **Split Java code into individual methods or classes (Improved)**
            snippets = java_split_code_snippets(file_code)  # ✅ Replaced re.split() with structured splitting
            summary["total_snippets"] += len(snippets)
            summary["total_lines"] += file_code.count('\n') + 1

            for line_num, snippet in enumerate(snippets, start=1):
                # ✅ **Check if snippet already analyzed**
                existing_snippet = JavaCodeSnippet.objects.filter(snippet=snippet).first()

                if existing_snippet:
                    print(f"✅ Found existing analysis for snippet in {file_name}...")

                    model_suggestion = existing_snippet.model_suggestion
                    ai_suggestion = existing_snippet.ai_suggestion

                    final_suggestion = f"{model_suggestion}\n\n💡Detail Suggestion:\n{ai_suggestion}"
                    category = categorize_suggestion(final_suggestion)
                    severity = determine_severity(final_suggestion)

                    summary["total_suggestions"] += 1
                    summary["categories"].setdefault(category, 0)
                    summary["categories"][category] += 1
                    summary["severity"][severity] += 1

                    suggestions.append({
                        "file_name": file_name,
                        "code": existing_snippet.snippet,
                        "suggestion": final_suggestion,
                        "category": category,
                        "severity": severity,
                        "line": line_num
                    })
                    continue  # ✅ Skip AI processing for existing snippets

                try:
                    print(f"🚀 Running AI analysis for snippet in {file_name}, Line {line_num}")

                    # ✅ **Generate Model-based Suggestion**
                    model_suggestion = java_generate_suggestion(snippet)

                    # ✅ **Generate AI-based Suggestion**
                    ai_suggestion = ai_code_analysis(snippet,)

                    final_suggestion = f"Suggestion:\n{model_suggestion}\n\nDetailed Analysis:\n{ai_suggestion}"

                    # ✅ **Categorize & determine severity**
                    category = categorize_suggestion(final_suggestion)
                    severity = determine_severity(final_suggestion)

                    summary["total_suggestions"] += 1
                    summary["categories"].setdefault(category, 0)
                    summary["categories"][category] += 1
                    summary["severity"][severity] += 1

                    # ✅ **Store in Java PostgreSQL tables**
                    JavaCodeSnippet.objects.create(
                        project=project,
                        snippet=snippet,
                        ai_suggestion=ai_suggestion,
                        model_suggestion=model_suggestion,
                    )

                    print(f"📌 Stored analysis for {file_name}, Line {line_num} in DB.")

                    suggestions.append({
                        "file_name": file_name,
                        "code": snippet,
                        "category": category,
                        "suggestion": final_suggestion,
                        "severity": severity,
                        "line": line_num
                    })

                except Exception as snippet_error:
                    print(f"❌ Error analyzing snippet in {file_name}, Line {line_num}: {str(snippet_error)}")
                    suggestions.append({
                        "file_name": file_name,
                        "code": snippet,
                        "suggestion": f"Error: {str(snippet_error)}",
                        "severity": "Low",
                        "line": line_num
                    })

        # ✅ **Perform Java Complexity Analysis**
        print("🔍 Performing Java Complexity Analysis...")
        summary["complexity_metrics"] = java_analyze_code_complexity(code_snippet)
        print(f"✅ Java Complexity Results: {summary['complexity_metrics']}")
        final_guideline = ai_generate_guideline(summary)

        # ✅ **Store the latest analysis in session for PDF generation**
        request.session["latest_summary"] = summary
        request.session["latest_suggestions"] = suggestions
        request.session["latest_guideline"] = final_guideline
        request.session.modified = True


    return render(
        request,
        'java_code_analysis.html',
        {'code': code_snippet, 'suggestions': suggestions, 'summary': summary, 'final_guideline': final_guideline}
    )




def java_analyze_code_complexity(code):
    """Analyze Java code complexity using various metrics."""
    print("🔍 Entering java_analyze_code_complexity function...")  # ✅ Debug
    guidelines = java_load_guidelines()

    loc, eloc = java_count_lines_of_code(code)
    num_classes, num_methods, avg_method_length = java_count_classes_and_methods(code)  # ✅ Corrected function
    # cyclomatic_complexity = java_calculate_cyclomatic_complexity(code)
    nesting_depth = java_calculate_nesting_depth(code)
    duplicate_percentage = java_count_duplicate_code_percentage(code)
    duplicate_code_details = java_find_duplicate_code(code)
    # **No `java_find_duplicate_code()` function exists, remove this line**
    # duplicate_code_details = java_find_duplicate_code(code) ❌ REMOVE THIS

    comment_density = java_calculate_comment_density(code)
    readability_score = java_calculate_readability_score(code)
    complexity_score = java_calculate_complexity_score(loc, num_methods, duplicate_percentage)

    result = {
        "lines_of_code": loc,
        "effective_lines_of_code": eloc,
        "num_classes": num_classes,  # ✅ Use classes instead of generic functions
        "num_methods": num_methods,
        "avg_method_length": avg_method_length,
        # "cyclomatic_complexity": cyclomatic_complexity,
        "nesting_depth": nesting_depth,
        "duplicate_code_percentage": duplicate_percentage,
        "duplicate_code_details": duplicate_code_details,
        "comment_density": comment_density,
        "readability_score": readability_score,
        "complexity_score": complexity_score,
        "rating": {
            "lines_of_code": java_categorize_value(loc, guidelines["lines_of_code"]),
            "comment_density": java_categorize_value(comment_density, guidelines["code_density"]),
            "method_length": java_categorize_value(avg_method_length, guidelines["function_length"]),
            "duplicate_code": java_categorize_value(duplicate_percentage, guidelines["duplicate_code"]),
            "num_methods": java_categorize_value(num_methods, guidelines["num_functions"]),
            # "cyclomatic_complexity": java_categorize_value(cyclomatic_complexity, guidelines["cyclomatic_complexity"]),
            "complexity_score": java_categorize_value(complexity_score, guidelines["complexity_score"]),
        }
    }

    print(f"✅ Java Complexity Analysis Output: {result}")  # ✅ Debug
    return result




@api_view(['GET', 'POST'])
def js_code_analyser(request):
    recommendations = {"files": {}, "pasted_code": {}}  # Initialize recommendations
    summary = {
        "total_vulnerabilities": 0,
        "categories": {},
        "files_analyzed": 0,
    }
    code = ""

    if request.method == 'POST':
        # Handle pasted code
        code = request.POST.get('code', '').strip()
        if code:
            try:
                analyzer = JavaScriptCodeAnalyzer(code)
                pasted_results = analyzer.generate_recommendations()
                recommendations["pasted_code"] = {}
                for rec in pasted_results:
                    line = rec.get('line', 'unknown')
                    recommendations["pasted_code"].setdefault(line, []).append({
                        'rule': rec.get('rule'),
                        'message': rec.get('message'),
                    })
                # Update summary
                for rec in pasted_results:
                    summary["total_vulnerabilities"] += 1
                    rule = rec.get('rule')
                    if rule:
                        summary["categories"].setdefault(rule, 0)
                        summary["categories"][rule] += 1
            except Exception as e:
                recommendations["pasted_code"] = {
                    "error": f"Error analyzing pasted code: {str(e)}"
                }

        # Handle uploaded files
        files = request.FILES.getlist('files')
        summary["files_analyzed"] = len(files)
        if files:
            for file in files:
                try:
                    content = file.read().decode('utf-8')  # Assuming UTF-8 encoding
                    analyzer = JavaScriptCodeAnalyzer(content)
                    file_results = analyzer.generate_recommendations()
                    recommendations["files"][file.name] = {}
                    for rec in file_results:
                        line = rec.get('line', 'unknown')
                        recommendations["files"][file.name].setdefault(line, []).append({
                            'rule': rec.get('rule'),
                            'message': rec.get('message'),
                        })
                    # Update summary
                    for rec in file_results:
                        summary["total_vulnerabilities"] += 1
                        rule = rec.get('rule')
                        if rule:
                            summary["categories"].setdefault(rule, 0)
                            summary["categories"][rule] += 1
                except Exception as e:
                    recommendations["files"][file.name] = {
                        "error": f"Error analyzing file {file.name}: {str(e)}"
                    }

        return render(
            request,
            'js_code_analyser.html',
            {'recommendations': recommendations, 'summary': summary, 'code': code}
        )

    return render(request, 'js_code_analyser.html', {'code': code})


@api_view(['GET', 'POST'])
def php_code_analyser(request):
    recommendations = {"pasted_code": {}, "files": {}}
    code = ""

    if request.method == 'POST':
        # Handle pasted code
        code = request.POST.get('code', '').strip()
        if code:
            try:
                analyzer = PHPCodeAnalyzer(code)
                pasted_results = analyzer.generate_recommendations()
                for rec in pasted_results:
                    line = rec.get('line', 'unknown')
                    recommendations["pasted_code"].setdefault(line, []).append({
                        'rule': rec.get('rule'),
                        'message': rec.get('message'),
                    })
            except Exception as e:
                recommendations["pasted_code"]["error"] = f"Error analyzing pasted code: {str(e)}"

        # Handle uploaded files
        files = request.FILES.getlist('files')
        if files:
            for file in files:
                try:
                    content = file.read().decode('utf-8')  # Assuming UTF-8 encoding
                    analyzer = PHPCodeAnalyzer(content)
                    file_results = analyzer.generate_recommendations()
                    recommendations["files"][file.name] = {}
                    for rec in file_results:
                        line = rec.get('line', 'unknown')
                        recommendations["files"][file.name].setdefault(line, []).append({
                            'rule': rec.get('rule'),
                            'message': rec.get('message'),
                        })
                except Exception as e:
                    recommendations["files"][file.name] = {
                        "error": f"Error analyzing file {file.name}: {str(e)}"
                    }

        return render(
            request,
            'php_code_analyser.html',
            {'recommendations': recommendations, 'code': code}
        )

    return render(request, 'php_code_analyser.html', {'code': code})


def normalize_and_validate_indentation(code_snippet):
    """
    Normalize and validate indentation in the code snippet to prevent parsing errors.
    """
    lines = code_snippet.splitlines()
    corrected_lines = []
    indent_stack = [0]  # Track indentation levels

    for index, line in enumerate(lines):
        stripped_line = line.strip()

        # Skip empty lines
        if not stripped_line:
            corrected_lines.append("")
            continue

        # Calculate current indentation level
        current_indent = len(line) - len(line.lstrip())

        # Check for invalid indentation
        if current_indent > indent_stack[-1]:
            # Ensure a block follows structures like def, if, for, etc.
            if corrected_lines and corrected_lines[-1].rstrip().endswith((':',)):
                indent_stack.append(current_indent)
            else:
                # Force valid indentation
                current_indent = indent_stack[-1]

        elif current_indent < indent_stack[-1]:
            while indent_stack and current_indent < indent_stack[-1]:
                indent_stack.pop()

        # Fix indentation if mismatched
        if current_indent != indent_stack[-1]:
            corrected_line = " " * indent_stack[-1] + stripped_line
        else:
            corrected_line = line

        corrected_lines.append(corrected_line)

    # Ensure every block with a colon has an indented body
    normalized_code = "\n".join(corrected_lines)
    fixed_code = ensure_blocks_have_bodies(normalized_code)

    return fixed_code


def ensure_blocks_have_bodies(code_snippet):
    """
    Add placeholder 'pass' statements to ensure blocks have valid bodies.
    """
    lines = code_snippet.splitlines()
    corrected_lines = []

    for index, line in enumerate(lines):
        stripped = line.strip()
        if stripped.endswith(":") and (index + 1 >= len(lines) or not lines[index + 1].strip()):
            corrected_lines.append(line)
            corrected_lines.append("    pass  # Placeholder for empty block")
        else:
            corrected_lines.append(line)

    return "\n".join(corrected_lines)


def split_code_snippets(code_snippet):
    """
    Split the input code snippet into individual Python functions or top-level blocks.
    """
    try:
        # Normalize and validate indentation
        normalized_code = normalize_and_validate_indentation(code_snippet)

        # Parse the normalized code into an Abstract Syntax Tree (AST)
        tree = ast.parse(normalized_code)
        snippets = []

        # Extract top-level nodes
        for node in tree.body:
            # Extract source code for each node
            if hasattr(ast, "get_source_segment"):
                snippet = ast.get_source_segment(normalized_code, node)
                print(f"Snippet: {snippet}")
            else:
                # Fallback: Use line numbers if available
                start_line = getattr(node, "lineno", None)
                end_line = getattr(node, "end_lineno", None)

                if start_line and end_line:
                    snippet_lines = normalized_code.splitlines()[start_line - 1:end_line]
                    snippet = "\n".join(snippet_lines)
                else:
                    # Fallback for cases where neither method works
                    snippet = ast.unparse(node) if hasattr(ast, "unparse") else ast.dump(node)

            snippets.append(snippet)

        return snippets
    except SyntaxError as e:
        print(f"Error parsing code snippets: {e}")
        return []  # Return an empty list if parsing fails







# Ensure API Key is loaded
openai.api_key = os.getenv("OPENAI_API_KEY")

# Ensure API Key is loaded
GITHUB_ACCESS_TOKEN = os.getenv("GITHUB_ACCESS_TOKEN")

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

try:
    response = client.models.list()
    print("✅ API Key is working. Available models:", response)
except Exception as e:
    print("❌ Invalid API Key or Quota Issue:", str(e))



# def fetch_github_files(repo_url):
#     """
#     Fetches code files from a GitHub repository (supports both public & private).
#     """
#     try:
#         # Extract repo details and branch
#         repo_parts = repo_url.replace("https://github.com/", "").split("/")
#         repo_owner, repo_name = repo_parts[0], repo_parts[1]
#         branch = repo_parts[3] if len(repo_parts) > 3 and repo_parts[2] == "tree" else "main"
#
#         # Fetch file list from the branch
#         api_url = f"https://api.github.com/repos/{repo_owner}/{repo_name}/git/trees/{branch}?recursive=1"
#         headers = {"Authorization": f"token {GITHUB_ACCESS_TOKEN}"}
#
#         response = requests.get(api_url, headers=headers)
#         if response.status_code != 200:
#             return None, f"Failed to fetch repository: {response.json().get('message', 'Unknown error')}"
#
#         files = response.json().get("tree", [])
#         code_files = []
#         allowed_extensions = {".py", ".js", ".java", ".cpp", ".cs", ".php"}
#
#         for file in files:
#             if file["type"] == "blob" and any(file["path"].endswith(ext) for ext in allowed_extensions):
#                 file_api_url = f"https://api.github.com/repos/{repo_owner}/{repo_name}/contents/{file['path']}?ref={branch}"
#                 file_response = requests.get(file_api_url, headers=headers)
#
#                 if file_response.status_code == 200:
#                     file_content = file_response.json().get("content", "")
#                     decoded_content = base64.b64decode(file_content).decode("utf-8")  # Decode base64 content
#
#                     code_files.append({"name": file["path"], "code": decoded_content})
#
#         return code_files, None
#
#     except Exception as e:
#         return None, f"Error fetching GitHub repository: {str(e)}"



def fetch_github_files(repo_url):
    """
    Fetches code files from a GitHub repository, including subfolders.
    """
    try:
        repo_parts = repo_url.replace("https://github.com/", "").split("/")
        repo_owner, repo_name = repo_parts[0], repo_parts[1]
        branch = repo_parts[3] if len(repo_parts) > 3 and repo_parts[2] == "tree" else "main"

        api_url = f"https://api.github.com/repos/{repo_owner}/{repo_name}/git/trees/{branch}?recursive=1"
        headers = {"Authorization": f"token {GITHUB_ACCESS_TOKEN}"}

        response = requests.get(api_url, headers=headers)
        if response.status_code != 200:
            return None, f"Failed to fetch repository: {response.json().get('message', 'Unknown error')}"

        files = response.json().get("tree", [])
        code_files = []
        allowed_extensions = {".py", ".js", ".java", ".cpp", ".cs", ".php"}

        for file in files:
            if file["type"] == "blob" and any(file["path"].endswith(ext) for ext in allowed_extensions):
                file_api_url = f"https://api.github.com/repos/{repo_owner}/{repo_name}/contents/{file['path']}?ref={branch}"
                file_response = requests.get(file_api_url, headers=headers)

                if file_response.status_code == 200:
                    file_content = file_response.json().get("content", "")
                    decoded_content = base64.b64decode(file_content).decode("utf-8")

                    # Store file name along with code content
                    code_files.append({"name": file["path"], "code": decoded_content})

        return code_files, None

    except Exception as e:
        return None, f"Error fetching GitHub repository: {str(e)}"



def ai_code_analysis(snippet):
    """
    Uses OpenAI's GPT-4o to analyze a given code snippet and provide structured suggestions.
    """
    try:
        response = client.chat.completions.create(
            model="gpt-3.5-turbo",  # Use "gpt-4o" for better results/ gpt-3.5-turbo
            messages=[
                {"role": "system",
                 "content": "You are an expert AI code reviewer. Follow this structure:\n"
                            "Issue Identified: Describe the issue concisely.\n"
                            "Why It's a Problem: Explain the consequences very shortly.\n"
                            "Recommended Fix: Provide a solution."},
                {"role": "user",
                 # "content": f"Analyze the following Python code and provide structured feedback:\n{snippet}"}
                 "content": f"Analyze the following code and provide structured feedback:\n{snippet}"}
            ],
            max_tokens=300,
            temperature=0.2  # Control randomness
        )

        # Extract AI response
        ai_response = response.choices[0].message.content

        # Format the response for HTML
        formatted_response = ai_response.replace("Issue Identified:", "<b>Issue Identified:</b>") \
            .replace("Why It's a Problem:", "<b>Why It's a Problem:</b>") \
            .replace("Recommended Fix:", "<b>Recommended Fix:</b>")

        return formatted_response  # ✅ Now formatted for HTML rendering
    except Exception as e:
        return f"Error generating AI analysis: {str(e)}"



def analyze_code_complexity(code):
    """Analyze code complexity using various metrics, including duplicate code detection."""
    print("🔍 Entering analyze_code_complexity function...")  # ✅ Debug
    guidelines = load_guidelines()

    loc, eloc = count_lines_of_code(code)
    num_functions, avg_function_length = count_functions_and_length(code)

    duplicate_percentage = count_duplicate_code_percentage(code)
    duplicate_code_details = find_duplicate_code(code)  # ✅ Get duplicate lines & locations

    comment_density = calculate_comment_density(code)
    readability_score = calculate_readability_score(code)
    complexity_score = calculate_complexity_score(loc, num_functions, duplicate_percentage)

    result = {
        "lines_of_code": loc,
        "effective_lines_of_code": eloc,
        "num_functions": num_functions,
        "avg_function_length": avg_function_length,
        "duplicate_code_percentage": duplicate_percentage,
        "duplicate_code_details": duplicate_code_details,  # ✅ Include duplicate details
        "comment_density": comment_density,
        "readability_score": readability_score,
        "complexity_score": complexity_score,
        "rating": {
            "lines_of_code": categorize_value(loc, guidelines["lines_of_code"]),
            "comment_density": categorize_value(comment_density, guidelines["code_density"]),
            "function_length": categorize_value(avg_function_length, guidelines["function_length"]),
            "duplicate_code": categorize_value(duplicate_percentage, guidelines["duplicate_code"]),
            "num_functions": categorize_value(num_functions, guidelines["num_functions"]),
            "complexity_score": categorize_value(complexity_score, guidelines["complexity_score"]),
        }
    }

    print(f"✅ Complexity Analysis Output: {result}")  # ✅ Debug: Print output
    return result



def ai_generate_guideline(summary):
    """
    Uses OpenAI to generate a final coding guideline for the developer based on the summary report.
    This includes best practices, security recommendations, maintainability tips, and efficiency improvements.
    """
    try:
        response = client.chat.completions.create(
            model="gpt-3.5-turbo",  # Upgrade for better contextual understanding
            messages=[
                {"role": "system",
                 "content": "You are an expert AI code reviewer. Based on the analysis summary, provide a concise final"
                            "guideline for the developer to follow. Keep it within 3-5 bullet points focusing on:\n"
                            "Use this format:\n\n"
                            "🚀 **Final Coding Guideline** 🚀\n\n"
                            "1️⃣ **Security Improvements:**\n"
                            "2️⃣ **Code Readability & Maintainability:**\n"
                            "3️⃣ **Performance Optimization:**\n"
                            "4️⃣  **Reference Links / Guidelines:**\n"
                 },
                {"role": "user",
                 "content": f"Here is the code analysis summary:\n{summary}\n\nGenerate a detailed final guideline for the developer to improve the overall code quality."}
            ],
            max_tokens=300,  # Increase token limit to provide detailed output
            temperature=0.2  # Keep responses structured and predictable
        )

        # Extract AI response
        guideline_response = response.choices[0].message.content

        # Format for HTML rendering
        formatted_guideline = guideline_response.replace("🚀 **Final Coding Guideline** 🚀", "<h3>🚀 Final Coding Guideline 🚀</h3>") \
            .replace("1️⃣ **Security Improvements:**", "<h4>🔒 Security Improvements</h4><ul>") \
            .replace("2️⃣ **Code Readability & Maintainability:**", "</ul><h4>📖 Code Readability & Maintainability</h4><ul>") \
            .replace("3️⃣ **Performance Optimization:**", "</ul><h4>⚡ Performance Optimization</h4><ul>") \
            .replace("4️⃣ **Reference Links / Guidelines:**", "</ul><h4>📚 Reference Guideline</h4><ul>")

        return formatted_guideline  # ✅ Now formatted for better HTML display
    except Exception as e:
        return f"Error generating final guideline: {str(e)}"


# def is_python_code(code):
#     """
#     Determines if the provided code is Python by checking syntax and keywords.
#     """
#     python_keywords = [
#         "import ", "def ", "class ", "lambda ", "yield ", "async ", "await ", "try:", "except ", "finally:", "with ",
#         "return ", "pass ", "break ", "continue ", "raise ", "global ", "nonlocal ", "assert ", "del ", "from ", "is ", "not "
#     ]
#
#     # Check if at least 2 Python-specific keywords exist
#     return sum(1 for kw in python_keywords if kw in code) >= 2


def is_python_code(code):
    """
    Checks if the provided code is valid Python code by attempting to parse it.
    """
    try:
        ast.parse(code)
        return True
    except SyntaxError:
        return False


@api_view(['GET', 'POST'])
def analyze_code_view(request):
    """
    Handles code analysis: Takes input, processes it, stores results in PostgreSQL, and retrieves past analysis if available.
    """
    suggestions = []
    summary = {
        "total_snippets": 0,
        "total_suggestions": 0,
        "total_lines": 0,
        "categories": {},
        "severity": {"Critical": 0, "Medium": 0, "Low": 0},
        "complexity_metrics": {},
    }
    code_snippet = ""
    final_guideline = ""
    all_code_snippets = []  # Store all code snippets from GitHub, uploaded files, and pasted code

    if request.method == 'POST':
        # ✅ Handle GitHub repository submission
        if request.content_type == 'application/json':
            data = json.loads(request.body)
            github_repo_url = data.get("github_url", "").strip()

            if github_repo_url:
                files, error = fetch_github_files(github_repo_url)
                if error:
                    return JsonResponse({"error": error})

                if files:
                    return JsonResponse({"files": files})  # ✅ Send GitHub files to frontend

        # ✅ Handle manually entered code & uploaded files
        code_snippet = request.POST.get('code', '').strip()
        project_name = request.POST.get('project_name', '').strip()

        # ✅ Auto-generate project name if none is provided
        if not project_name:
            project_name = f"Project_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            print(f"🆕 Auto-generated project name: {project_name}")

        # ✅ Check if project already exists, else create it
        project, _ = Project.objects.get_or_create(name=project_name)

        # ✅ Fetch uploaded files
        uploaded_files = request.FILES.getlist('files')

        # ✅ Process uploaded files
        if uploaded_files:
            for uploaded_file in uploaded_files:
                file_name = uploaded_file.name
                file_content = uploaded_file.read().decode('utf-8')
                all_code_snippets.append({"name": file_name, "code": file_content})
        else:
            # Handle code pasted directly in the editor
            if code_snippet:
                all_code_snippets.append({"name": "Pasted Code", "code": code_snippet})

        if not all_code_snippets:
            return render(request, 'analyze_code.html', {
                "code": "", "suggestions": [], "summary": summary, "final_guideline": "",
                "error": "No code provided for analysis."
            })

        for file_data in all_code_snippets:
            file_name = file_data["name"]
            file_code = file_data["code"]

            # ✅ **Check if the code is Python before analysis**
            if not is_python_code(file_code):
                return render(request, 'analyze_code.html', {
                    "code": file_code,
                    "suggestions": [],
                    "summary": summary,
                    "final_guideline": "",
                    "error": f" Error: The uploaded file `{file_name}` is not valid Python code."
                })

            snippets = split_code_snippets(file_code)
            summary["total_snippets"] += len(snippets)
            summary["total_lines"] += file_code.count('\n') + 1

            for line_num, snippet in enumerate(snippets, start=1):
                # ✅ **Check if this snippet has already been analyzed**
                existing_snippet = CodeSnippet.objects.filter(snippet=snippet).first()

                if existing_snippet:
                    print(f"✅ Found existing analysis for snippet in {file_name}...")

                    # ✅ **Retrieve stored model and AI suggestions**
                    model_suggestion = existing_snippet.model_suggestion
                    ai_suggestion = existing_snippet.ai_suggestion

                    # ✅ **Combine previous model & AI suggestions for display**
                    final_suggestion = (
                        f"\n{model_suggestion}\n\n"
                        f"💡Detail Suggestion :\n{ai_suggestion}"
                    )

                    # ✅ **Ensure categories & severity are updated from DB suggestions**
                    category = categorize_suggestion(final_suggestion)
                    severity = determine_severity(final_suggestion)

                    summary["total_suggestions"] += 1
                    summary["categories"].setdefault(category, 0)
                    summary["categories"][category] += 1
                    summary["severity"][severity] += 1

                    # ✅ **Store combined suggestions in the results**
                    suggestions.append({
                        "file_name": file_name,  # ✅ Associate file name
                        "code": existing_snippet.snippet,
                        "suggestion": final_suggestion,  # ✅ Use combined suggestion
                        "category": category,
                        "severity": severity,
                        "line": line_num
                    })
                    continue  # ✅ Skip AI & Model processing for existing snippets

                try:
                    print(f"🚀 Running AI analysis for snippet in {file_name}, Line {line_num}")

                    # ✅ **Perform AI-based Model Analysis**
                    t5_suggestion = call_flask_model(snippet)  # Calls Flask API
                    gpt_suggestion = ai_code_analysis(snippet)

                    # ✅ **Combine AI-generated suggestions**
                    final_suggestion = f"Suggestion:\n{t5_suggestion}\n\nDetailed Analysis:\n{gpt_suggestion}"

                    # ✅ **Categorize and determine severity**
                    category = categorize_suggestion(final_suggestion)
                    severity = determine_severity(final_suggestion)

                    # ✅ **Update summary**
                    summary["total_suggestions"] += 1
                    summary["categories"].setdefault(category, 0)
                    summary["categories"][category] += 1
                    summary["severity"][severity] += 1

                    # ✅ **Store the snippet analysis in PostgreSQL**
                    CodeSnippet.objects.create(
                        project=project,
                        snippet=snippet,
                        ai_suggestion=gpt_suggestion,
                        model_suggestion=t5_suggestion,
                    )

                    print(f"📌 Stored analysis for {file_name}, Line {line_num} in DB.")

                    # ✅ **Store the suggestion in results**
                    suggestions.append({
                        "file_name": file_name,  # ✅ Track filename here
                        "code": snippet,
                        "category": category,
                        "suggestion": final_suggestion,  # ✅ Store full AI + Model suggestion
                        "severity": severity,
                        "line": line_num
                    })

                except Exception as snippet_error:
                    print(f"❌ Error analyzing snippet in {file_name}, Line {line_num}: {str(snippet_error)}")
                    suggestions.append({
                        "file_name": file_name,  # ✅ Ensure errors also show filename
                        "code": snippet,
                        "suggestion": f"Error: {str(snippet_error)}",
                        "severity": "Low",
                        "line": line_num
                    })

        # ✅ **Perform Complexity Analysis**
        print("🔍 Performing Complexity Analysis...")
        summary["complexity_metrics"] = analyze_code_complexity(code_snippet)
        print(f"✅ Complexity Results: {summary['complexity_metrics']}")

        # ✅ **Store the latest analysis in session for PDF generation**
        request.session["latest_summary"] = summary
        request.session["latest_suggestions"] = suggestions
        request.session.modified = True

        # ✅ **Generate Developer Guideline**
        final_guideline = ai_generate_guideline(summary)

    return render(
        request,
        'analyze_code.html',
        {'code': code_snippet, 'suggestions': suggestions, 'summary': summary, 'final_guideline': final_guideline}
    )




# @api_view(['GET', 'POST'])
# def analyze_code_view(request):
#     """
#     Handles code analysis: Takes input, processes it, stores results in PostgreSQL, and retrieves past analysis if available.
#     """
#     suggestions = []
#     summary = {
#         "total_snippets": 0,
#         "total_suggestions": 0,
#         "total_lines": 0,
#         "categories": {},
#         "severity": {"Critical": 0, "Medium": 0, "Low": 0},
#         "complexity_metrics": {},
#     }
#     code_snippet = ""
#     final_guideline = ""
#
#     if request.method == 'POST':
#         # ✅ Handle GitHub repository submission
#         if request.content_type == 'application/json':
#             data = json.loads(request.body)
#             github_repo_url = data.get("github_url", "").strip()
#
#             if github_repo_url:
#                 files, error = fetch_github_files(github_repo_url)
#                 if error:
#                     return JsonResponse({"error": error})
#
#                 if files:
#                     return JsonResponse({"files": files})  # ✅ Return fetched GitHub files
#
#         # ✅ Handle normal code submission
#         code_snippet = request.POST.get('code', '').strip()
#         project_name = request.POST.get('project_name', '').strip()
#
#         # ✅ Auto-generate project name if none is provided
#         if not project_name:
#             project_name = f"Project_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
#             print(f"🆕 Auto-generated project name: {project_name}")
#
#         # ✅ Check if project already exists, else create it
#         project, _ = Project.objects.get_or_create(name=project_name)
#
#         if not code_snippet:
#             return render(request, 'analyze_code.html', {
#                 "code": "", "suggestions": [], "summary": summary, "final_guideline": "",
#                 "error": "No code provided for analysis."
#             })
#
#         try:
#             print(f"📥 Received Code Snippet: {code_snippet[:100]}...")  # ✅ DEBUG: Verify input
#             snippets = split_code_snippets(code_snippet)
#             summary["total_snippets"] = len(snippets)
#             summary["total_lines"] = code_snippet.count('\n') + 1
#
#             for line_num, snippet in enumerate(snippets, start=1):
#                 # ✅ **Check if this snippet has already been analyzed**
#                 existing_snippet = CodeSnippet.objects.filter(snippet=snippet).first()
#
#                 if existing_snippet:
#                     print(f"✅ Found existing analysis for snippet: {snippet[:50]}...")
#
#                     # ✅ **Ensure categories & severity are updated from DB suggestions**
#                     category = categorize_suggestion(existing_snippet.model_suggestion)
#                     severity = determine_severity(existing_snippet.model_suggestion)
#
#                     summary["total_suggestions"] += 1
#                     summary["categories"].setdefault(category, 0)
#                     summary["categories"][category] += 1
#                     summary["severity"][severity] += 1
#
#                     suggestions.append({
#                         "code": existing_snippet.snippet,
#                         "suggestion": existing_snippet.model_suggestion,
#                         "category": category,
#                         "severity": severity,
#                         "line": line_num
#                     })
#                     continue  # Skip AI & Model processing for existing snippets
#
#                 try:
#                     print(f"🚀 No previous analysis found. Running AI and Model analysis for snippet {line_num}")
#
#                     # ✅ **Perform AI-based T5 Model Analysis**
#                     # inputs = tokenizer(
#                     #     snippet, truncation=True, padding="max_length", max_length=512, return_tensors="pt"
#                     # )
#                     # inputs = {key: value.to(device) for key, value in inputs.items()}
#                     # model.eval()
#                     #
#                     # with torch.no_grad():
#                     #     outputs = model.generate(
#                     #         inputs["input_ids"], max_length=256, num_beams=5, early_stopping=True
#                     #     )
#                     #     t5_suggestion = tokenizer.decode(outputs[0], skip_special_tokens=True)
#                     print(f"🚀 Calling Flask model for snippet {line_num}...")
#                     t5_suggestion = call_flask_model(snippet)  # Calls Flask API
#
#                     # ✅ **Generate AI-based GPT analysis**
#                     gpt_suggestion = ai_code_analysis(snippet)
#
#                     # ✅ **Combine AI-generated suggestions**
#                     final_suggestion = f"Suggestion:\n{t5_suggestion}\n\nDetailed Analysis:\n{gpt_suggestion}"
#
#                     # ✅ **Categorize and determine severity**
#                     category = categorize_suggestion(final_suggestion)
#                     severity = determine_severity(final_suggestion)
#
#                     # ✅ **Update summary**
#                     summary["total_suggestions"] += 1
#                     summary["categories"].setdefault(category, 0)
#                     summary["categories"][category] += 1
#                     summary["severity"][severity] += 1
#
#                     # ✅ **Store the snippet analysis in PostgreSQL**
#                     CodeSnippet.objects.create(
#                         project=project,
#                         snippet=snippet,
#                         ai_suggestion=gpt_suggestion,
#                         model_suggestion=t5_suggestion,
#                     )
#
#                     print(f"📌 Stored analysis for new snippet {line_num} in DB.")
#
#                     # ✅ **Store the suggestion in results**
#                     suggestions.append({
#                         "code": snippet,
#                         "category": category,
#                         "suggestion": final_suggestion,
#                         "severity": severity,
#                         "line": line_num
#                     })
#
#                 except Exception as snippet_error:
#                     print(f"❌ Error analyzing snippet {line_num}: {str(snippet_error)}")
#                     suggestions.append({
#                         "code": snippet,
#                         "suggestion": f"Error: {str(snippet_error)}",
#                         "severity": "Low",
#                         "line": line_num
#                     })
#
#             # ✅ **Perform Complexity Analysis**
#             print("🔍 Performing Complexity Analysis...")
#             summary["complexity_metrics"] = analyze_code_complexity(code_snippet)
#             print(f"✅ Complexity Results: {summary['complexity_metrics']}")
#             print(f"✅ Categories Assigned: {summary['categories']}")
#
#             # ✅ **Store the latest analysis in session for PDF generation**
#             request.session["latest_summary"] = summary  # ✅ Store summary
#             request.session["latest_suggestions"] = suggestions  # ✅ Store suggestions
#             request.session.modified = True  # Ensure session updates
#
#             # ✅ **Generate Developer Guideline**
#             final_guideline = ai_generate_guideline(summary)
#
#         except Exception as e:
#             print(f"❌ Critical Error in analysis: {str(e)}")
#             suggestions.append({
#                 "code": "",
#                 "suggestion": f"Error analyzing code: {str(e)}",
#                 "severity": "Critical",
#                 "line": 0
#             })
#
#     return render(
#         request,
#         'analyze_code.html',
#         {'code': code_snippet, 'suggestions': suggestions, 'summary': summary, 'final_guideline': final_guideline}
#     )
#






# def compare_trend(previous_value, current_value):
#     """Compares two values and returns an indicator if it improved, worsened, or stayed the same."""
#     if current_value < previous_value:
#         return f"✅ Improved ({previous_value} → {current_value})"
#     elif current_value > previous_value:
#         return f"❌ Increased ({previous_value} → {current_value})"
#     else:
#         return f"➖ No Change ({previous_value})"





def export_excel(request):
    """Generate and download the Excel report with detailed insights."""
    summary = request.session.get("latest_summary", {})  # Get latest analysis
    suggestions = request.session.get("latest_suggestions", [])  # Get latest suggestions

    # Create Excel response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="code_analysis_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

    # Create Workbook and Worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Code Analysis Report"

    # Define Styles
    bold_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Gray for headers
    table_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")  # Light gray for tables
    wrap_alignment = Alignment(wrap_text=True, vertical="top")  # ✅ Enable wrapping
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )  # ✅ Add cell borders

    # ✅ Add Report Title and Timestamp
    ws.append(["Code Analysis Report"])
    ws.append(["Generated On", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append([])

    # Apply Bold Style to Title and Date
    ws["A1"].font = title_font
    ws["A2"].font = bold_font
    ws["B2"].font = bold_font

    # ✅ Add Summary Metrics
    ws.append(["Metric", "Value"])
    ws.append(["Total Code Segments Analyzed", summary.get("total_snippets", 0)])
    ws.append(["Total Suggestions", summary.get("total_suggestions", 0)])
    ws.append(["Total Issues Identified", summary.get("total_suggestions", 0)])  # ✅ Added Total Issues
    ws.append(["Total Lines of Code", summary.get("total_lines", 0)])
    ws.append(["Number of Functions", summary.get("complexity_metrics", {}).get("num_functions", "N/A")])
    ws.append(["Duplicate Code Percentage", summary.get("complexity_metrics", {}).get("duplicate_code_percentage", "N/A")])
    ws.append(["Average Function Length", summary.get("complexity_metrics", {}).get("avg_function_length", "N/A")])
    ws.append(["Comment Density", summary.get("complexity_metrics", {}).get("comment_density", "N/A")])
    ws.append(["Complexity Score", summary.get("complexity_metrics", {}).get("complexity_score", "N/A")])
    ws.append(["Readability Score", summary.get("complexity_metrics", {}).get("readability_score", "N/A")])
    ws.append(["Duplicate Code Percentage", summary.get("complexity_metrics", {}).get("duplicate_code_percentage", "N/A")])
    ws.append([])

    # ✅ Add Severity Levels
    ws.append(["Severity Level", "Count"])
    ws.append(["Critical Issues", summary.get("severity", {}).get("Critical", 0)])
    ws.append(["Medium Issues", summary.get("severity", {}).get("Medium", 0)])
    ws.append(["Low Issues", summary.get("severity", {}).get("Low", 0)])
    ws.append([])

    # ✅ Add Category Breakdown
    ws.append(["Category", "Count"])
    for category, count in summary.get("categories", {}).items():
        ws.append([category, count])
    ws.append([])

    # ✅ Add Code Snippets with Issues & Suggestions (Table Formatting)
    header_row = ["Code Snippet", "Suggested Fix", "Category", "Severity", "Line Number"]
    ws.append(header_row)

    # ✅ Apply Styling to Headers
    for col_num, cell in enumerate(ws[ws.max_row], start=1):
        cell.font = bold_font
        cell.fill = header_fill  # Apply background color
        cell.alignment = wrap_alignment
        cell.border = thin_border  # ✅ Apply borders to header

    # ✅ Process & Add Data to Table
    if suggestions:
        for suggestion in suggestions:
            row_data = [
                suggestion.get("code", "").replace("\n", " "),  # ✅ Ensure inline format
                suggestion.get("suggestion", "N/A"),  # ✅ Full Suggested Fix
                suggestion.get("category", "N/A"),
                suggestion.get("severity", "N/A"),
                suggestion.get("line", "N/A"),
            ]
            ws.append(row_data)

        # ✅ Enable text wrapping and apply borders to data rows
        for row in ws.iter_rows(min_row=ws.max_row - len(suggestions) + 1, max_row=ws.max_row):
            for cell in row:
                cell.alignment = wrap_alignment
                cell.border = thin_border  # ✅ Add borders for each cell
                cell.fill = table_fill  # ✅ Apply background color for readability

    else:
        ws.append(["No analyzed code snippets available."])

    # ✅ Adjust Column Widths Dynamically for Better Readability
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # Get column letter (A, B, C, etc.)

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        # ✅ Adjust width dynamically while setting a maximum width
        if col_letter == "A":  # ✅ Code Snippet Column
            adjusted_width = min(max_length + 10, 80)  # ✅ Increase for better readability
        elif col_letter == "B":  # ✅ Suggested Fix Column
            adjusted_width = min(max_length + 10, 80)  # ✅ Allow space for suggestions
        else:
            adjusted_width = min(max_length + 5, 30)  # ✅ Keep other columns readable

        ws.column_dimensions[col_letter].width = adjusted_width

    # ✅ Save workbook to response
    wb.save(response)
    return response







def categorize_suggestion(suggestion):
    """Categorize a suggestion based on its content using meaningful categories."""
    suggestion_lower = suggestion.lower()

    # Security-related issues
    if any(term in suggestion_lower for term in [
        "sql injection", "command injection", "hardcoded secret",
        "authentication", "authorization", "xss", "csrf", "security",
        "data exposure", "encryption", "password hashing", "secret key"
    ]):
        return "Security Vulnerabilities"

    # Performance-related improvements
    elif any(term in suggestion_lower for term in [
        "optimize", "performance", "efficiency", "redundant code",
        "slow database query", "memory leak", "loop optimization",
        "batch processing", "unnecessary computation"
    ]):
        return "Performance Enhancements"

    # Code Quality & Maintainability
    elif any(term in suggestion_lower for term in [
        "refactor", "code quality", "readability", "maintainability",
        "magic numbers", "naming conventions", "global variables",
        "long function", "long class", "duplicate code", "function too complex",
        "dead code", "improve clarity"
    ]):
        return "Code Readability and Maintainability"

    # Resource & Memory Management Issues
    elif any(term in suggestion_lower for term in [
        "resource leak", "memory management", "insecure file handling",
        "open file without closing", "socket not closed",
        "unclosed connection", "improper resource release"
    ]):
        return "Resource and Memory Management"

    # Exception Handling Issues
    elif any(term in suggestion_lower for term in [
        "empty exception handling", "broad exception handling",
        "unhandled exception", "try except pass", "error handling",
        "catch generic exception", "silent failure"
    ]):
        return "Exception Handling Issues"

    # Deprecated & Compatibility Issues
    elif any(term in suggestion_lower for term in [
        "deprecated function", "deprecated library", "legacy code",
        "backward compatibility", "unsupported method",
        "modern alternatives"
    ]):
        return "Deprecated and Compatibility Issues"

    # Object-Oriented Programming (OOP) Best Practices
    elif any(term in suggestion_lower for term in [
        "inheritance misuse", "misuse of polymorphism", "encapsulation violation",
        "tight coupling", "improper abstraction", "large class",
        "single responsibility principle violation"
    ]):
        return "Object-Oriented Design Issues"

    # API & Cloud Security Issues
    elif any(term in suggestion_lower for term in [
        "exposed api key", "weak api authentication",
        "rate limiting missing", "caching missing", "api security",
        "insecure endpoint", "unvalidated input"
    ]):
        return "API and Cloud Security Issues"

    # Business Logic Vulnerabilities
    elif any(term in suggestion_lower for term in [
        "flawed business logic", "logic flaw", "invalid transaction",
        "unauthorized access", "insufficient validation", "race condition"
    ]):
        return "Business Logic Vulnerabilities"

    # Compliance & Standard Violations
    elif any(term in suggestion_lower for term in [
        "violates coding standards", "non-compliant", "does not follow",
        "missing documentation", "inconsistent style", "linting issues",
        "coding standard violation"
    ]):
        return "Compliance and Standard Violations"

    # API Design Issues
    elif any(term in suggestion_lower for term in [
        "bad api design", "poor rest implementation", "missing status codes",
        "improper json response", "error response missing", "misuse of http verbs"
    ]):
        return "API Design and Best Practices"

    # Multithreading & Concurrency Issues
    elif any(term in suggestion_lower for term in [
        "race condition", "deadlock", "thread safety", "improper synchronization",
        "mutex missing", "concurrent modification"
    ]):
        return "Multithreading & Concurrency Issues"

    return "General Issues"


def determine_severity(suggestion):
    """Determine the severity level of a suggestion."""
    suggestion_lower = suggestion.lower()

    # Critical Severity
    if any(term in suggestion_lower for term in [
        "sql injection", "command injection", "hardcoded secret", "secret key",
        "weak cryptography", "directory traversal", "resource leak"
    ]):
        return "Critical"

    # Medium Severity
    if any(term in suggestion_lower for term in [
        "should fix", "deep nesting", "global variables",
        "deprecated libraries", "variable shadowing", "magic numbers",
        "unreachable code", "empty exception handling"
    ]):
        return "Medium"

    # Low Severity
    if any(term in suggestion_lower for term in [
        "consistent naming", "consistent return types",
        "unused imports", "excessive comments", "consistent whitespace"
    ]):
        return "Low"

    # Default to Low if no matches
    return "Low"


# def split_code_snippets(code_snippet):
#     """
#     Split the input code snippet into individual Python functions or top-level blocks.
#     """
#     try:
#         # Parse the input code into an Abstract Syntax Tree (AST)
#         tree = ast.parse(code_snippet)
#         snippets = []
#
#         for node in tree.body:  # Iterate over top-level statements
#             # Get the source code for each node
#             snippet = ast.unparse(node) if hasattr(ast, 'unparse') else ast.dump(node)
#             snippets.append(snippet)
#
#         return snippets
#     except Exception as e:
#         print(f"Error parsing code snippets: {e}")
#         return []  # Return an empty list if parsing fails
#
#
#
# @api_view(['GET', 'POST'])
# def analyze_code_view(request):
#     """
#     Analyze the given code snippet(s) for vulnerabilities using the trained model.
#     """
#     suggestions = []
#     code_snippet = ""
#
#     if request.method == 'POST':
#         # Get the code snippet(s) from the form
#         code_snippet = request.POST.get('code', '').strip()
#         if not code_snippet:
#             suggestions.append({"code": "", "suggestion": "Please provide valid code snippet(s)."})
#         else:
#             try:
#                 # Debug: Log the received code snippet(s)
#                 print(f"Received code snippet(s):\n{code_snippet}")
#
#                 # Split code into individual snippets
#                 snippets = split_code_snippets(code_snippet)
#                 if not snippets:
#                     suggestions.append({"code": "", "suggestion": "Error: Unable to parse code snippets."})
#                 else:
#                     # Debug: Log parsed snippets
#                     print("Parsed code snippets:", snippets)
#
#                     # Load the trained model and tokenizer
#                     model_path = "./models/custom_seq2seq_model"  # Adjust path if needed
#                     tokenizer = AutoTokenizer.from_pretrained(model_path)
#                     model = T5ForConditionalGeneration.from_pretrained(model_path)
#                     device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
#                     model.to(device)
#
#                     # Process each snippet separately
#                     for snippet in snippets:
#                         # Tokenize the individual snippet
#                         inputs = tokenizer(
#                             snippet,
#                             truncation=True,
#                             padding="max_length",
#                             max_length=512,
#                             return_tensors="pt"
#                         )
#                         inputs = {key: value.to(device) for key, value in inputs.items()}
#
#                         # Generate suggestion for the snippet
#                         model.eval()
#                         with torch.no_grad():
#                             outputs = model.generate(
#                                 inputs["input_ids"],
#                                 max_length=256,  # Increased max_length for longer outputs
#                                 num_beams=5,     # Using beam search for better suggestions
#                                 early_stopping=True
#                             )
#
#                             # Decode the suggestion
#                             suggestion = tokenizer.decode(outputs[0], skip_special_tokens=True)
#
#                         # Append the result as a dictionary
#                         suggestions.append({"code": snippet, "suggestion": suggestion})
#
#             except Exception as e:
#                 # Handle errors during model inference
#                 suggestions.append({"code": "", "suggestion": f"Error analyzing the code: {str(e)}"})
#
#     return render(
#         request,
#         'analyze_code.html',
#         {'code': code_snippet, 'suggestions': suggestions}
#     )


def detect_defects_view(request):
    model_path = "./models/defect_detection_model"
    model = RobertaForSequenceClassification.from_pretrained(model_path)
    tokenizer = RobertaTokenizer.from_pretrained(model_path)

    if request.method == "POST":
        try:
            # Handle pasted code
            code_snippet = request.POST.get("code_snippet", "").strip()
            if not code_snippet:
                return JsonResponse({"error": "No code snippet provided."}, status=400)

            inputs = tokenizer(
                code_snippet, return_tensors="pt", truncation=True, padding="max_length", max_length=512
            )
            outputs = model(**inputs)
            prediction = torch.argmax(outputs.logits).item()

            return JsonResponse({"defect_detected": bool(prediction)})

        except Exception as e:
            return JsonResponse({"error": f"Detection failed: {str(e)}"}, status=500)

    return render(request, "detect_defects.html")


# @api_view(['GET', 'POST'])
# def java_code_analysis(request):
#     if request.method == 'POST':
#         try:
#             code = request.POST.get('code', '')
#             if not code:
#                 return JsonResponse({'error': 'No code provided'}, status=400)
#
#             analyser = JavaCodeAnalyzer(code)
#             recommendations = analyser.generate_recommendations()
#             return JsonResponse({'recommendations': recommendations})
#
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)
#
#     return render(request, 'java_code_analysis.html')


# @api_view(['GET', 'POST'])
# def upload_python_files(request):
#     if request.method == 'POST':
#         files = request.FILES.getlist('files')
#         if not files:
#             return JsonResponse({'error': 'No files uploaded'}, status=400)
#
#         results = {}
#         for file in files:
#             try:
#                 content = file.read().decode('utf-8')  # Assuming UTF-8 encoding
#                 analyzer = PythonCodeAnalyzer(content)
#                 recommendations = analyzer.generate_recommendations()
#                 results[file.name] = recommendations
#             except Exception as e:
#                 results[file.name] = f"Error analyzing file: {str(e)}"
#
#         return JsonResponse({'results': results})
#
#     return render(request, 'upload_python.html')


# @api_view(['GET', 'POST'])
# def upload_java_files(request):
#     if request.method == 'POST':
#         files = request.FILES.getlist('files')
#         if not files:
#             return JsonResponse({'error': 'No files uploaded'}, status=400)
#
#         results = {}
#         for file in files:
#             try:
#                 content = file.read().decode('utf-8')  # Assuming UTF-8 encoding
#                 analyzer = JavaCodeAnalyzer(content)
#                 recommendations = analyzer.generate_recommendations()
#                 results[file.name] = recommendations
#             except Exception as e:
#                 results[file.name] = f"Error analyzing file: {str(e)}"
#
#         return JsonResponse({'results': results})
#
#     return render(request, 'upload_java.html')

@api_view(['GET', 'POST'])
def drink_list(request, format=None):
    if request.method == 'GET':
        drinks = Drink.objects.all()
        serializer = DrinkSerializer(drinks, many=True)
        return Response(serializer.data)
    if request.method == 'POST':
        serializer = DrinkSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)


@api_view(['GET', 'PUT', 'DELETE'])
def drink_detail(request, id, format=None):
    Drink.objects.get(pk=id)

    try:
        drink = Drink.objects.get(pk=id)
    except Drink.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = DrinkSerializer(drink)
        return Response(serializer.data)
    elif request.method == 'PUT':
        serializer = DrinkSerializer(drink, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    elif request.method == 'DELETE':
        drink.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# @api_view(['POST', 'GET'])
# def calculate_complexity_line_by_line(request):
#     if request.method == 'POST':
#         code = request.data.get('code')
#         if not code:
#             return Response({'error': 'Code is required'}, status=status.HTTP_400_BAD_REQUEST)
#
#         complexity = calculate_code_complexity_line_by_line(code)
#         return Response(complexity, status=status.HTTP_200_OK)



@api_view(['GET', 'POST'])
def calculate_complexity_line_by_line(request):
    if request.method == 'POST':
        code = request.data.get('code')
        if not code:
            return Response({'error': 'Code is required'}, status=status.HTTP_400_BAD_REQUEST)

        # Call your function to calculate complexity
        result= calculate_code_complexity_line_by_line(code)
        complexity = result['line_complexities']
        cbo = result['cbo']

        # Render the result in the template
        return render(request, 'complexityA_table.html', {'complexities': complexity, 'cbo':cbo})

    # If GET request, just show the form
    return render(request, 'complexityA_form.html')

def get_thresholds():
    thresholds_file = os.path.join(settings.BASE_DIR, 'media', 'threshold4.json')
    if os.path.exists(thresholds_file):
        with open(thresholds_file, 'r') as json_file:
            thresholds = json.load(json_file)
        return thresholds
    # Default values if the file doesn't exist
    return {'threshold_low': 10, 'threshold_medium': 20}


@api_view(['GET', 'POST'])
def calculate_complexity_multiple_java_files(request):
    if request.method == 'POST':
        try:
            # Expecting multiple Java files in the request
            files = request.FILES.getlist('files')
            if not files:
                return Response({'error': 'No files uploaded'}, status=status.HTTP_400_BAD_REQUEST)

            file_contents = {file.name: file.read().decode('utf-8') for file in files}

            # Load thresholds from the JSON file
            thresholds = get_thresholds()
            print("thresholds<<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>>>>>>>>>>", thresholds)
            threshold_low = thresholds.get('threshold_low', 10)
            # threshold_medium = thresholds.get('threshold_medium', 20)
            threshold_high = thresholds.get('threshold_high', 50)

            # Calculate complexity for multiple Java files
            result, cbo_predictions = calculate_code_complexity_multiple_files(file_contents)

            complexities, cbo_summary = [], []
            results_table = PrettyTable()
            results_table.field_names = ["Filename", "Line Number", "Line", "Size", "Tokens",
                                         "Control Structure Complexity", "Nesting Weight",
                                         "Inheritance Weight", "Compound Condition Weight",
                                         "Try-Catch Weight", "Thread Weight", "CBO", "Total Complexity"]

            mp_cbo_table = PrettyTable()
            mp_cbo_table.field_names = ["Filename", "CBO", "MPC"]
            saved_files = []

            # Process complexity results
            for filename, file_data in result.items():
                complexity_data = file_data.get('complexity_data', [])
                cbo, mpc = file_data.get('cbo', 0), file_data.get('mpc', 0)
                method_complexities = file_data.get('method_complexities', {})
                recommendations = file_data.get('recommendation', [])
                pie_chart_path = file_data.get('pie_chart_path', '')
                total_wcc = file_data.get('total_wcc', 0)
                bar_charts = file_data.get('bar_charts', {})

                java_file, created = JavaFile.objects.update_or_create(
                    filename=filename,
                    defaults={'total_wcc': file_data.get('total_wcc', 0)}
                )
                saved_files.append({'filename': filename, 'total_wcc': java_file.total_wcc})

                for line_data in complexity_data:
                    if len(line_data) == 12:
                        results_table.add_row([filename] + line_data)
                    else:
                        print(f"Skipping malformed data for {filename}: {line_data}")

                # Categorize methods based on complexity
                categorized_methods = []
                for method_name, method_data in method_complexities.items():
                    if isinstance(method_data, dict):
                        total_complexity = method_data.get('total_complexity', 0)
                        if total_complexity <= threshold_low:
                            category = 'Low'
                        elif threshold_low <= total_complexity <= threshold_high:
                            category = 'Medium'
                        else:
                            category = 'High'

                        categorized_methods.append({
                            **method_data,
                            'category': category,
                            'method_name': method_name,
                            'bar_chart': bar_charts.get(method_name, '')
                        })
                    else:
                        print(f"Unexpected format in method_data: {method_data}")

                complexities.append({
                    'filename': filename,
                    'complexity_data': complexity_data,
                    'cbo': cbo,
                    'mpc': mpc,
                    'method_complexities': categorized_methods,
                    'recommendations': recommendations,
                    'pie_chart_path': pie_chart_path,
                    'total_wcc': total_wcc
                })

            # Extract CBO Predictions & Recommendations
            for filename, prediction_data in cbo_predictions.items():
                cbo_summary.append({
                    'filename': filename,
                    'prediction': prediction_data.get('prediction', 'Unknown'),
                    'recommendations': prediction_data.get('recommendations', [])
                })

            # Return JSON if requested
            if request.headers.get('Accept') == 'application/json':
                return Response({
                    'complexities': complexities,
                    'cbo_predictions': cbo_summary
                }, status=status.HTTP_200_OK)

            # Render template for web-based UI
            return render(request, 'complexity_table.html',
                          {'complexities': complexities, 'cbo_predictions': cbo_summary})


        except JavaSyntaxError as e:
            return Response({
                'error': 'Java Syntax Error detected. Please correct your Java code.',
                'details': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({
                'error': 'An unexpected error occurred.',
                'details': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # Render form for GET requests
    return render(request, 'complexity_form.html')


@api_view(['GET', 'POST'])
def calculate_complexity_line_by_line_csharp(request):
    if request.method == 'POST':
        code = request.data.get('code')
        if not code:
            return Response({'error': 'Code is required'}, status=status.HTTP_400_BAD_REQUEST)

        # Call the function to calculate complexity line-by-line for a C# file
        result = calculate_code_complexity_line_by_line(code)
        complexity = result['line_complexities']
        cbo = result['cbo']

        # Render the result in the template
        return render(request, 'complexityB_table.html', {'complexities': complexity, 'cbo': cbo})

    # If GET request, just show the form
    return render(request, 'complexityB_form.html')


@api_view(['GET', 'POST'])
def calculate_complexity_multiple_csharp_files(request):
    if request.method == 'POST':
        # Expecting multiple Java files in the request
        files = request.FILES.getlist('files')
        if not files:
            return Response({'error': 'No files uploaded'}, status=status.HTTP_400_BAD_REQUEST)

        file_contents = {}
        for file in files:
            # Read the content of each file
            content = file.read().decode('utf-8')  # Assuming the files are UTF-8 encoded
            file_contents[file.name] = content

        # Load thresholds from the JSON file
        thresholds = get_thresholds()
        print("thresholds.....]]]]]]]]]]]]]]]]]]]", thresholds)
        threshold_low = thresholds.get('threshold_low', 10)
        threshold_medium = thresholds.get('threshold_medium', 20)
        threshold_high = thresholds.get('threshold_high', 50)

        # Call your function to calculate complexity for multiple files
        result, cbo_predictions = calculate_code_complexity_multiple_files_csharp(file_contents)

        # Prepare a list to store complexities for each file
        complexities = []
        cbo_summary = []

        # Create a table for the response
        results_table = PrettyTable()
        results_table.field_names = ["Filename", "Line Number", "Line", "Size", "Tokens",
                                     "Control Structure Complexity", "Nesting Weight",
                                     "Inheritance Weight", "Compound Condition Weight",
                                     "Try-Catch Weight", "Thread Weight", "CBO", "Total Complexity"]

        # Prepare another table for displaying MPC and CBO values
        mp_cbo_table = PrettyTable()
        mp_cbo_table.field_names = ["Filename", "CBO", "MPC"]

        # Collect complexities for each file
        for filename, file_data in result.items():
            complexity_data = file_data['complexity_data']
            cbo = file_data['cbo']
            mpc = file_data['mpc']
            method_complexities = file_data['method_complexities']
            recommendations = file_data['recommendation']
            pie_chart_path = file_data['pie_chart_path']
            total_wcc = file_data['total_wcc']
            bar_charts = file_data.get('bar_charts', {})

            print("method_complexities.....]]]]]]]]]]{{{{{{{{{{{{{{{{{{{{{{{{{{{{{{{{{{{{{", method_complexities)

            for line_data in complexity_data:
                results_table.add_row([filename] + line_data)  # Now line_data has 9 values

            # # Categorize each method based on total complexity
            categorized_methods = []
            for method_name, method_data in method_complexities.items():
                if isinstance(method_data, dict):
                    total_complexity = method_data.get('total_complexity', 0)

                    print("total_complexity<<<<<<<<<<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>>>>>>>>>>>>>>>", total_complexity)

                    if total_complexity <= threshold_low:
                        category = 'Low'
                    elif threshold_low <= total_complexity <= threshold_high:
                        category = 'Medium'
                    else:
                        category = 'High'

                    # Append category to the method data
                    # categorized_method = method_data.copy()
                    # categorized_method['category'] = category
                    # categorized_method['method_name'] = method_name
                    # categorized_methods.append(categorized_method)

                    categorized_methods.append({
                        **method_data,
                        'category': category,
                        'method_name': method_name,
                        'bar_chart': bar_charts.get(method_name, '')
                    })
                else:
                    print(f"Unexpected format in method_data: {method_data}")

                # print("categorized_method", categorized_method)
            complexities.append({
                'filename': filename,
                'complexity_data': complexity_data,
                'cbo': cbo,
                'mpc': mpc,
                'method_complexities': categorized_methods,
                'recommendations': recommendations,
                'pie_chart_path': pie_chart_path,
                'total_wcc': total_wcc
            })

        # Log the result table for debugging or reference
        # print(results_table)
        for filename, prediction_data in cbo_predictions.items():
            cbo_summary.append({
                'filename': filename,
                'prediction': prediction_data['prediction'],
                'recommendations': prediction_data['recommendations']
            })
        if request.headers.get('Accept') == 'application/json':
            return Response({
                'complexities': complexities,
                'cbo_predictions': cbo_summary
            }, status=status.HTTP_200_OK)

        # Instead of returning a JSON response, render the template and pass complexities
        return render(request, 'complexityC_table.html', {'complexities': complexities, 'cbo_predictions': cbo_summary})

    # If GET request, just show the form
    return render(request, 'complexityC_form.html')


@api_view(['GET', 'POST'])
def calculate_complexity(request):
    if request.method == 'POST':
        uploaded_file = request.FILES['file']

        if uploaded_file.name.endswith('.csv'):
            # Define the media directory path
            media_dir = os.path.join(settings.BASE_DIR, 'media')
            os.makedirs(media_dir, exist_ok=True)  # Create the directory if it doesn't exist

            # Save the uploaded file
            file_path = os.path.join(media_dir, uploaded_file.name)
            with open(file_path, 'wb+') as destination:
                for chunk in uploaded_file.chunks():
                    destination.write(chunk)

            # Load the CSV into a DataFrame
            data = pd.read_csv(file_path)

            # Columns to analyze
            target_columns = ['Complexity', 'Maintainability', 'Readability']

            # Initialize correlation results
            pearson_results, spearman_results = {}, {}

            # Perform correlation analysis
            for column in target_columns:
                pearson_corr, pearson_p_value = pearsonr(data['WCC'], data[column])
                spearman_corr, spearman_p_value = spearmanr(data['WCC'], data[column])

                pearson_results[column] = {
                    'correlation': round(pearson_corr, 4),
                    'p_value': round(pearson_p_value, 4)
                }
                spearman_results[column] = {
                    'correlation': round(spearman_corr, 4),
                    'p_value': round(spearman_p_value, 4)
                }

            # Generate Correlation Matrix
            correlation_matrix = data[['WCC', 'Complexity', 'Maintainability', 'Readability']].corr()
            sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm', center=0)
            heatmap_path = os.path.join(media_dir, 'heatmap.png')
            plt.title('Correlation Matrix')
            plt.savefig(heatmap_path)
            plt.close()

            # Clustering for WCC thresholds
            wcc_values = np.array([101, 191, 205, 246, 291, 293, 298, 344, 346, 382,
                                   170, 61, 108, 183, 204, 153, 305, 270, 137, 201,
                                   345, 73, 114, 153]).reshape(-1, 1)

            plt.figure(figsize=(10, 6))
            sns.histplot(wcc_values, bins=10, kde=True, color='blue')

            # Labels and title
            plt.xlabel("WCC Values")
            plt.ylabel("Frequency")
            plt.title("WCC Distribution Histogram")
            plt.grid(True)
            histogram_path = os.path.join(media_dir, 'wcc_distribution.png')
            plt.savefig(histogram_path)
            plt.close()

            kmeans = KMeans(n_clusters=2, random_state=0, n_init=10).fit(wcc_values)
            data['Cluster_Level'] = kmeans.predict(data[['WCC']])

            # Map clusters to categories based on cluster centers
            cluster_centers = sorted(kmeans.cluster_centers_.flatten())
            low_center, high_center = cluster_centers[0], cluster_centers[1]
            data['Cluster_Level'] = data['Cluster_Level'].map({
                0: 'Low Complexity',
                1: 'High Complexity'
            })

            # Save thresholds to JSON
            thresholds = {
                'threshold_low': round(low_center, 2),
                # 'threshold_medium': round(medium_center, 2),
                'threshold_high': round(high_center, 2)
            }

            # Path to save the thresholds
            threshold_file_path = os.path.join(media_dir, 'threshold4.json')
            with open(threshold_file_path, 'w') as json_file:
                json.dump(thresholds, json_file)

            # Scatter plots for each metric
            scatter_plots = {}
            for column in target_columns:
                plt.figure(figsize=(10, 6))
                sns.regplot(x=data['WCC'], y=data[column], scatter=True, ci=None, line_kws={'color': 'red'})
                plt.title(f'WCC vs {column}')
                plt.xlabel('WCC')
                plt.ylabel(column)
                scatter_plot_path = os.path.join(media_dir, f'scatter_{column}.png')
                plt.savefig(scatter_plot_path)
                plt.close()
                scatter_plots[column] = scatter_plot_path

                # Boxplot to show WCC thresholds with KMeans cluster thresholds
                plt.figure(figsize=(10, 6))
                sns.boxplot(x='Cluster_Level', y='WCC', data=data)

                # Add horizontal lines for KMeans thresholds
                plt.axhline(y=low_center, color='red', linestyle='--', label=f'Low Threshold ({round(low_center, 2)})')
                # plt.axhline(y=medium_center, color='orange', linestyle='--',
                #             label=f'Medium Threshold ({round(medium_center, 2)})')
                plt.axhline(y=high_center, color='green', linestyle='--',
                            label=f'High Threshold ({round(high_center, 2)})')

                plt.title('WCC by Complexity Level with KMeans Thresholds')
                plt.ylabel('WCC')
                plt.xlabel('Cluster Level')
                plt.legend()

                # Save the updated boxplot
                boxplot_path = os.path.join(media_dir, 'boxplot_with_thresholds.png')
                plt.savefig(boxplot_path)
                plt.close()

            # Prepare context for the result page
            context = {
                'pearson_results': pearson_results,
                'spearman_results': spearman_results,
                'scatter_plots': scatter_plots,
                'low_center': round(low_center, 2),
                # 'medium_center': round(medium_center, 2),
                'high_center': round(high_center, 2),
                'kmeans_centers': cluster_centers,
                'plot_path': 'scatter_Complexity.png',
                'heatmap_path': 'heatmap.png',
            }

            # Render the result page
            return render(request, 'analysis_result.html', context)

        else:
            return HttpResponse("Invalid file format. Please upload a CSV file.")

    return render(request, 'upload.html')


def guidelines_view(request):
    guidelines = [
        {
            "title": "Weight Due to Type of Control Structures (Wc)",
            "description": "Control structures influence code complexity by introducing decision paths.",
            "table": [
                {"structure": "Sequential Statements", "weight": 0},
                {"structure": "Branching (if, else, else if)", "weight": 1},
                {"structure": "Loops (for, while, do-while)", "weight": 2},
                {"structure": "Switch statement with n cases", "weight": "n"},
            ],
            "examples": {
                "Java": """
                    public class ControlExample {
                        public static void main(String[] args) {
                            int num = 10;

                            // Branching (if-else) - Weight: 1
                            if (num > 0) {
                                System.out.println("Positive number");
                            } else {
                                System.out.println("Negative number");
                            }

                            // Loop (for) - Weight: 2
                            for (int i = 0; i < 5; i++) {
                                System.out.println(i);
                            }

                            // Switch statement (3 cases) - Weight: 3
                            switch (num) {
                                case 5: System.out.println("Five"); break;
                                case 10: System.out.println("Ten"); break;
                                default: System.out.println("Other"); break;
                            }
                        }
                    }
                    """,
                "Csharp": """
                    using System;

                    class ControlExample {
                        static void Main() {
                            int num = 10;

                            // Branching (if-else) - Weight: 1
                            if (num > 0)
                                Console.WriteLine("Positive number");
                            else
                                Console.WriteLine("Negative number");

                            // Loop (while) - Weight: 2
                            int i = 0;
                            while (i < 5) {
                                Console.WriteLine(i);
                                i++;
                            }

                            // Switch statement (3 cases) - Weight: 3
                            switch (num) {
                                case 5: Console.WriteLine("Five"); break;
                                case 10: Console.WriteLine("Ten"); break;
                                default: Console.WriteLine("Other"); break;
                            }
                        }
                    }
                    """
            }
        },
        {
            "title": "Weight Due to Nesting Level of Control Structures (Wn)",
            "description": "Nesting increases complexity as deeply nested statements are harder to understand.",
            "table": [
                {"structure": "Outermost statements", "weight": 1},
                {"structure": "Second level", "weight": 2},
                {"structure": "nth level", "weight": "n"},
            ],
            "examples": {
                "Java": """
    public class NestingExample {
        public static void main(String[] args) {
            int num = 5;

            if (num > 0) {  // Level 1 - Weight: 1
                for (int i = 0; i < 3; i++) {  // Level 2 - Weight: 2
                    while (num > 0) {  // Level 3 - Weight: 3
                        System.out.println(num);
                        num--;
                    }
                }
            }
        }
    }
                """,
                "Csharp": """
    using System;

    class NestingExample {
        static void Main() {
            int num = 5;

            if (num > 0) { // Level 1 - Weight: 1
                for (int i = 0; i < 3; i++) { // Level 2 - Weight: 2
                    while (num > 0) { // Level 3 - Weight: 3
                        Console.WriteLine(num);
                        num--;
                    }
                }
            }
        }
    }
                """
            }
        },
        {
            "title": "Weight Due to Inheritance Level of Statements (Wi)",
            "description": "Statements in derived classes increase complexity as they depend on parent classes.",
            "table": [
                {"structure": "Base class", "weight": 1},
                {"structure": "First-level derived class", "weight": 2},
                {"structure": "nth level derived class", "weight": "n"},
            ],
            "examples": {
                "Java": """
    class Animal {  // Base class - Weight: 1
        void makeSound() {
            System.out.println("Animal sound"); 
        }
    }

    class Dog extends Animal { // First-level derived - Weight: 2
        @Override
        void makeSound() { 
            System.out.println("Bark");
        }
    }

    class Puppy extends Dog { // Second-level derived - Weight: 3
        @Override
        void makeSound() { 
            System.out.println("Small Bark");
        }
    }
                """,
                "Csharp": """
    using System;

    class Animal {  // Base class - Weight: 1
        public virtual void MakeSound() {
            Console.WriteLine("Animal sound");
        }
    }

    class Dog : Animal { // First-level derived - Weight: 2
        public override void MakeSound() { 
            Console.WriteLine("Bark");
        }
    }

    class Puppy : Dog { // Second-level derived - Weight: 3
        public override void MakeSound() { 
            Console.WriteLine("Small Bark");
        }
    }
                """
            }
        },
        {
            "title": "Coupling Between Object Classes (Wcbo)",
            "description": "CBO measures the degree of dependency between classes. High coupling increases complexity and reduces maintainability.",
            "table": [
                {"structure": "Loose coupling (e.g., dependency injection via constructor/setter)", "weight": 1},
                {
                    "structure": "Tight coupling (e.g., static method call, static variable usage, direct object instantiation)",
                    "weight": 3},
            ],
            "examples": {
                "Java": """
class Engine {
    public int rpm;
    static void start() { 
        System.out.println("Engine started");
    }
}

class Car {
    private Engine engine;

    public void setEngine(Engine engine) { // Setter Injection
        this.engine = engine; // Loose Coupling - Weight: 1
    }

    void drive() {
        engine.start();
        Engine.rpm = 1000; // Static variable usage (Tight Coupling - Weight: 3)
        System.out.println("Car is driving");
    }
}

class CarTight {
    void drive() {
        Engine.start(); // Static method call (Tight Coupling - Weight: 3)
        System.out.println("Car is using an engine of type: " + Engine.type); // Static variable usage (Tight Coupling - Weight: 3)
    }
}

                    """,
                "Csharp": """
            class Engine {

    public static void Start() { // Static method call
        Console.WriteLine("Engine started");
    }
}

class Car {
    private Engine _engine;

    public void SetEngine(Engine engine) { // Setter Injection
        _engine = engine; // Loose Coupling - Weight: 1
    }

    public void Drive() {
        _engine.Start();
        Console.WriteLine("Car is driving");
    }
}

class CarTight {
    public void Drive() {
        Engine.Start(); // Static method call (Tight Coupling - Weight: 3)
        Console.WriteLine($"Car is using an engine of type: {Engine.Type}"); // Static variable usage (Tight Coupling - Weight: 3)
    }
}
                    """
            }
        },
        {
            "title": "Weight Due to Try-Catch-Finally Blocks (Wtc)",
            "description": "Try-Catch-Finally structures influence code complexity by adding exception-handling paths. The weight is determined based on the nesting depth and control structure type.",
            "table": [
                {"structure": "try-catch",
                 "guideline": "Assigned weight based on nesting depth. Deeper nesting increases complexity.",
                 "weight": "1 (Level 1), 2 (Level 2), 3 (Level 3), 4 (Level 4+)"},
                {"structure": "finally",
                 "guideline": "Always executes, adding a mandatory execution path. Assigned a fixed weight.",
                 "weight": 1}
            ],
            "examples": {
                "Java": """
            public class ExceptionHandlingExample {
                public static void main(String[] args) {
                    try {
                        int result = 10 / 0; 
                    } catch (ArithmeticException e) { // Catch at Level 1 (Weight: 1)
                        System.out.println("Division by zero!"); 
                    }

                    try {
                        try {
                            int[] arr = {1, 2, 3};
                            System.out.println(arr[5]); 
                        } catch (ArrayIndexOutOfBoundsException e) { // Catch at Level 2 (Weight: 2)
                            System.out.println("Index out of bounds!"); 
                        }
                    } catch (Exception e) { // Catch at Level 1 (Weight: 1)
                        System.out.println("Generic exception!"); 
                    } finally { // Finally block (Weight: 1)
                        System.out.println("Execution finished."); 
                    }
                }
            }
                    """,
                "Csharp": """
            using System;

            class ExceptionHandlingExample {
                static void Main() {
                    try {
                        int result = 10 / 0; 
                    } catch (DivideByZeroException e) {  // Catch at Level 1 (Weight: 1)
                        Console.WriteLine("Division by zero!");
                    }

                    try {
                        try {
                            int[] arr = {1, 2, 3};
                            Console.WriteLine(arr[5]); 
                        } catch (IndexOutOfRangeException e) { // Catch at Level 2 (Weight: 2)
                            Console.WriteLine("Index out of bounds!"); 
                        }
                    } catch (Exception e) { // Catch at Level 1 (Weight: 1)
                        Console.WriteLine("Generic exception!"); 
                    } finally { // Finally block (Weight: 1)
                        Console.WriteLine("Execution finished."); 
                    }
                }
            }
                    """
            }
        },
        {
            "title": "Weight Due to Compound Conditional Statements(Wcc)",
            "description": "Logical operators increase complexity.",
            "table": [
                {"structure": "Simple condition", "weight": 1},
                {"structure": "Compound condition with n logical operators", "weight": "n"},
            ],
            "examples": {
                "Java": """
    if (age > 18) {  // Weight: 1
        System.out.println("Adult");
    }

    if (age > 18 && country.equals("USA")) {  // Weight: 2
        System.out.println("Eligible voter in the USA");
    }
                """,
                "Csharp": """
    if (age > 18) { // Weight: 1
        Console.WriteLine("Adult");
    }

    if (age > 18 && country == "USA") { // Weight: 2
        Console.WriteLine("Eligible voter in the USA");
    }
                """
            }
        },
        {
            "title": "Weight Due to Threads (Wth)",
            "description": "Multi-threading increases complexity. Thread creation and synchronization mechanisms contribute to concurrency overhead.",
            "table": [
                {
                    "structure": "Simple thread creation",
                    "guideline": "Creating a new thread increases complexity.\n\nJava:\n```java\nThread t1 = new Thread(() -> System.out.println(\"Thread running\"));\nt1.start();\n```\n\nC#:\n```csharp\nThread t1 = new Thread(() => Console.WriteLine(\"Thread running\"));\nt1.Start();\n```",
                    "weight": 2
                },
                {
                    "structure": "Basic synchronized block",
                    "guideline": "Using synchronized blocks to protect shared resources.\n\nJava:\n```java\nsynchronized (this) {\n    System.out.println(\"Synchronized block\");\n}\n```\n\nC#:\n```csharp\nobject lockObj = new object();\nlock (lockObj) {\n    Console.WriteLine(\"Synchronized block\");\n}\n```",
                    "weight": 3
                },
                {
                    "structure": "Nested synchronized block",
                    "guideline": "Synchronization inside another synchronized block increases complexity.\n\nJava:\n```java\nsynchronized (this) {\n    synchronized (this) {\n        System.out.println(\"Nested synchronized block\");\n    }\n}\n```\n\nC#:\n```csharp\nlock (lockObj) {\n    lock (lockObj) {\n        Console.WriteLine(\"Nested synchronized block\");\n    }\n}\n```",
                    "weight": 4
                },
                {
                    "structure": "Method-level synchronization",
                    "guideline": "Declaring a method as synchronized increases complexity significantly.\n\nJava:\n```java\npublic synchronized void syncMethod() {\n    System.out.println(\"Synchronized method\");\n}\n```\n\nC#:\n```csharp\nprivate static readonly object _lock = new object();\npublic void SyncMethod() {\n    lock (_lock) {\n        Console.WriteLine(\"Synchronized method\");\n    }\n}\n```",
                    "weight": 5
                }
            ],
            "examples": {
                "Java": "```java\nclass ThreadExample {\n    public static void main(String[] args) {\n        // Simple Thread Creation - Weight: 2\n        Thread t1 = new Thread(() -> System.out.println(\"Thread 1 running\"));\n        t1.start();\n\n        // Basic Synchronized Block - Weight: 3\n        synchronized (this) {\n            System.out.println(\"Synchronized block\");\n        }\n\n        // Nested Synchronized Block - Weight: 4\n        synchronized (this) {\n            synchronized (this) {\n                System.out.println(\"Nested synchronized block\");\n            }\n        }\n\n        // Method-Level Synchronization - Weight: 5\n        new ThreadExample().syncMethod();\n    }\n\n    public synchronized void syncMethod() {\n        System.out.println(\"Synchronized method\");\n    }\n}\n```",
                "Csharp": "```csharp\nusing System;\nusing System.Threading;\n\nclass ThreadExample {\n    public static void Main() {\n        // Simple Thread Creation - Weight: 2\n        Thread t1 = new Thread(() => Console.WriteLine(\"Thread 1 running\"));\n        t1.Start();\n\n        // Basic Synchronized Block - Weight: 3\n        object lockObj = new object();\n        lock (lockObj) {\n            Console.WriteLine(\"Synchronized block\");\n        }\n\n        // Nested Synchronized Block - Weight: 4\n        lock (lockObj) {\n            lock (lockObj) {\n                Console.WriteLine(\"Nested synchronized block\");\n            }\n        }\n\n        // Method-Level Synchronization - Weight: 5\n        new ThreadExample().SyncMethod();\n    }\n\n    private static readonly object _lock = new object();\n    public void SyncMethod() {\n        lock (_lock) {\n            Console.WriteLine(\"Synchronized method\");\n        }\n    }\n}\n```"
            }
        }

    ]

    return render(request, 'guidelines.html', {"guidelines": guidelines})
